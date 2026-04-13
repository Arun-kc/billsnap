"""
Export service — generates CSV and Excel files from bills for a given month.
"""

import csv
import io
from datetime import date

import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from ..models.bill import Bill
from ..models.line_item import LineItem

_BILL_HEADERS = [
    "Bill Date", "Bill Number", "Vendor Name", "Vendor GSTIN",
    "Document Type", "Category", "Taxable Amount",
    "CGST", "SGST", "IGST", "Total Amount", "Verified", "Notes",
]

_LINE_ITEM_HEADERS = [
    "Bill Number", "Item Name", "HSN Code", "Qty", "Unit",
    "Unit Price", "Total Price", "GST Rate %",
]


async def _fetch_bills(db: AsyncSession, user_id, month: str) -> list[Bill]:
    y, m = int(month[:4]), int(month[5:7])
    result = await db.execute(
        select(Bill)
        .where(Bill.user_id == user_id)
        .where(Bill.bill_date >= date(y, m, 1))
        .where(Bill.bill_date < date(y + (m == 12), (m % 12) + 1, 1))
        .options(selectinload(Bill.line_items))
        .order_by(Bill.bill_date, Bill.bill_number)
    )
    return list(result.scalars().all())


def _bill_row(bill: Bill) -> list:
    return [
        bill.bill_date.isoformat() if bill.bill_date else "",
        bill.bill_number or "",
        bill.vendor_name or "",
        bill.vendor_gstin or "",
        bill.document_type,
        bill.category or "",
        float(bill.taxable_amount) if bill.taxable_amount else "",
        float(bill.cgst_amount),
        float(bill.sgst_amount),
        float(bill.igst_amount),
        float(bill.total_amount) if bill.total_amount else "",
        "Yes" if bill.is_verified else "No",
        bill.user_notes or "",
    ]


async def generate_csv(db: AsyncSession, user_id, month: str) -> bytes:
    bills = await _fetch_bills(db, user_id, month)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(_BILL_HEADERS)
    for bill in bills:
        writer.writerow(_bill_row(bill))
    return buf.getvalue().encode("utf-8-sig")  # utf-8-sig for Excel compatibility


async def generate_excel(db: AsyncSession, user_id, month: str) -> bytes:
    bills = await _fetch_bills(db, user_id, month)

    wb = openpyxl.Workbook()
    # --- Bills sheet ---
    ws_bills = wb.active
    ws_bills.title = "Bills"
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    ws_bills.append(_BILL_HEADERS)
    for cell in ws_bills[1]:
        cell.fill = header_fill
        cell.font = header_font

    for bill in bills:
        ws_bills.append(_bill_row(bill))

    # Totals row
    if bills:
        ws_bills.append([
            "", "", "", "", "", "TOTAL",
            sum(float(b.taxable_amount or 0) for b in bills),
            sum(float(b.cgst_amount) for b in bills),
            sum(float(b.sgst_amount) for b in bills),
            sum(float(b.igst_amount) for b in bills),
            sum(float(b.total_amount or 0) for b in bills),
            "", "",
        ])
        for cell in ws_bills[ws_bills.max_row]:
            cell.font = Font(bold=True)

    # Auto-width
    for col in ws_bills.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws_bills.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    # --- Line Items sheet ---
    ws_items = wb.create_sheet("Line Items")
    ws_items.append(_LINE_ITEM_HEADERS)
    for cell in ws_items[1]:
        cell.fill = header_fill
        cell.font = header_font

    for bill in bills:
        for item in bill.line_items:
            ws_items.append([
                bill.bill_number or "",
                item.item_name or "",
                item.hsn_code or "",
                float(item.quantity) if item.quantity else "",
                item.unit or "",
                float(item.unit_price) if item.unit_price else "",
                float(item.total_price) if item.total_price else "",
                float(item.gst_rate) if item.gst_rate else "",
            ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
