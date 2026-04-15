"""
Unit tests for app.services.export_service.

Tests generate real CSV and Excel bytes using mocked bill data,
so no database or filesystem is required.
"""

import csv
import io
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest

from app.services.export_service import (
    _BILL_HEADERS,
    _LINE_ITEM_HEADERS,
    generate_csv,
    generate_excel,
)
from tests.conftest import USER_ID, make_bill, make_line_item, make_mock_db


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_bills_with_items():
    bill = make_bill()
    bill.line_items = [make_line_item(bill_id=bill.id)]
    return [bill]


def _make_empty_bills():
    return []


# ---------------------------------------------------------------------------
# generate_csv
# ---------------------------------------------------------------------------

class TestGenerateCsv:
    async def test_csv_has_header_row(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _make_bills_with_items()
            db = make_mock_db()

            csv_bytes = await generate_csv(db, USER_ID, "2026-04")

        text = csv_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        assert rows[0] == _BILL_HEADERS

    async def test_csv_data_row_matches_bill(self):
        bills = _make_bills_with_items()
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = bills
            db = make_mock_db()

            csv_bytes = await generate_csv(db, USER_ID, "2026-04")

        text = csv_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        # Row index 1 = first data row
        assert len(rows) == 2  # header + 1 bill
        data_row = rows[1]
        assert data_row[2] == "Kerala Electricals"  # Vendor Name column
        assert data_row[10] == "1180.0"  # Total Amount column

    async def test_csv_empty_month_has_header_only(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _make_empty_bills()
            db = make_mock_db()

            csv_bytes = await generate_csv(db, USER_ID, "2026-01")

        text = csv_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)

        assert len(rows) == 1  # header only
        assert rows[0] == _BILL_HEADERS

    async def test_csv_verified_column(self):
        bill = make_bill(is_verified=True)
        bill.line_items = []
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = [bill]
            db = make_mock_db()

            csv_bytes = await generate_csv(db, USER_ID, "2026-04")

        text = csv_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert rows[1][11] == "Yes"  # Verified column

    async def test_csv_unverified_column(self):
        bill = make_bill(is_verified=False)
        bill.line_items = []
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = [bill]
            db = make_mock_db()

            csv_bytes = await generate_csv(db, USER_ID, "2026-04")

        text = csv_bytes.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        assert rows[1][11] == "No"

    async def test_csv_utf8_bom_prefix(self):
        """CSV should start with UTF-8 BOM for Excel compatibility."""
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = []
            db = make_mock_db()
            csv_bytes = await generate_csv(db, USER_ID, "2026-04")

        assert csv_bytes[:3] == b"\xef\xbb\xbf"


# ---------------------------------------------------------------------------
# generate_excel
# ---------------------------------------------------------------------------

class TestGenerateExcel:
    async def test_excel_has_bills_sheet(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _make_bills_with_items()
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert "Bills" in wb.sheetnames

    async def test_excel_has_line_items_sheet(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = _make_bills_with_items()
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        assert "Line Items" in wb.sheetnames

    async def test_excel_bills_header_row(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = []
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Bills"]
        headers = [cell.value for cell in ws[1]]
        assert headers == _BILL_HEADERS

    async def test_excel_line_items_header_row(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = []
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Line Items"]
        headers = [cell.value for cell in ws[1]]
        assert headers == _LINE_ITEM_HEADERS

    async def test_excel_data_row_present(self):
        bills = _make_bills_with_items()
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = bills
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Bills"]
        # Row 2 = first data row
        vendor_cell = ws.cell(row=2, column=3).value  # Vendor Name is column 3
        assert vendor_cell == "Kerala Electricals"

    async def test_excel_totals_row_appended(self):
        """When bills exist, a TOTAL row should be appended."""
        bills = _make_bills_with_items()
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = bills
            db = make_mock_db()

            xlsx_bytes = await generate_excel(db, USER_ID, "2026-04")

        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["Bills"]
        # header row + 1 bill + 1 total row = 3 rows
        assert ws.max_row == 3
        total_cell = ws.cell(row=3, column=6).value
        assert total_cell == "TOTAL"

    async def test_excel_returns_bytes(self):
        with patch("app.services.export_service._fetch_bills", new_callable=AsyncMock) as mock_fetch:
            mock_fetch.return_value = []
            db = make_mock_db()

            result = await generate_excel(db, USER_ID, "2026-04")

        assert isinstance(result, bytes)
        assert len(result) > 0
